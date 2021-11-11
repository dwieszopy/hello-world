import pandas as pd
from openpyxl import load_workbook
from copy import copy

path = "/tmp/Book1-in.xlsx"

wb = load_workbook(filename=path)
ws = wb['Sheet1']
merged = copy(ws.merged_cells.ranges)

for cell_range in merged:
    ws.unmerge_cells(str(cell_range))
    for row in ws[str(cell_range)]:
        for cell in row:
            cell.value = ws[str(cell_range).split(':')[0]].value

data = ws.values
columns = next(ws.values)[0:]
data = ws.values
df = pd.DataFrame(data, columns=columns).fillna("")

# df["col2"] = df.groupby(["col1","col4"])["col2"].transform(lambda x: ",".join(sorted(list(set(x)))).lstrip(","))
# df = df.drop_duplicates()
# df
df = df.groupby(["col1","col4"]).aggregate({"col2" : lambda x: ",".join(sorted(list(set(x)))).lstrip(",")})
df
